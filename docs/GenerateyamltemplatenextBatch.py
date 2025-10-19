#!/usr/bin/env python3
"""
Generate YAML template for next batch of hospitals from Excel file
Usage: python generate_hospital_yaml.py
"""

import openpyxl
from datetime import datetime

# Configuration
EXCEL_FILE = 'LeadershipURLYAMLNotes.xlsx'
OUTPUT_FILE = 'next_batch_template.yaml'
SHEET_NAME = 'LookupTypeFAC'
BATCH_SIZE = 30

def format_fac(fac_num):
    """Format FAC number as 3-digit string"""
    return f"{int(fac_num):03d}"

def clean_name(name):
    """Clean and format hospital name"""
    if not name:
        return ""
    return str(name).strip().upper()

def generate_yaml_entry(fac, name):
    """Generate a YAML entry for one hospital"""
    fac_formatted = format_fac(fac)
    name_clean = clean_name(name)
    
    return f'''
 - FAC: "{fac_formatted}"
   name: "{name_clean}"
   url: ""  # ← REQUIRED: Add leadership page URL
   pattern: "combined_h2"  # ← UPDATE: Change to match actual pattern
   expected_executives: 5  # ← UPDATE: Change based on website
   html_structure:
     combined_element: ""  # ← CUSTOMIZABLE 
     separator: ""  # ← CUSTOMIZABLE
     notes: ""  # ← OPTIONAL: Add any special notes
   status: "needs_configuration"  # ← UPDATE: Change to 'needs_testing' once URL added
'''

def main():
    print("=" * 70)
    print("HOSPITAL YAML TEMPLATE GENERATOR")
    print("=" * 70)
    print()
    
    # Load workbook
    print(f"Loading {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # Get the sheet
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found!")
        print(f"Available sheets: {', '.join(wb.sheetnames)}")
        return
    
    ws = wb[SHEET_NAME]
    print(f"Reading from sheet: {SHEET_NAME}")
    
    # Find header row (assuming row 1 has headers)
    headers = {}
    for idx, cell in enumerate(ws[1], 1):
        if cell.value:
            headers[cell.value] = idx
    
    print(f"Found columns: {', '.join(headers.keys())}")
    
    # Check required columns
    required = ['FAC', 'Hospital', 'done']
    missing = [col for col in required if col not in headers]
    if missing:
        print(f"ERROR: Missing required columns: {', '.join(missing)}")
        return
    
    fac_col = headers['FAC']
    name_col = headers['Hospital']
    done_col = headers['done']
    
    # Read hospitals
    hospitals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        fac = row[fac_col - 1]
        name = row[name_col - 1]
        done = row[done_col - 1]
        
        # Skip if no FAC
        if not fac:
            continue
            
        # Skip if marked done
        done_str = str(done).lower().strip() if done else ""
        if done_str == 'y':
            continue
        
        hospitals.append({'FAC': fac, 'Hospital': name})
    
    print(f"\nTotal hospitals in file: {ws.max_row - 1}")
    print(f"Hospitals NOT marked done: {len(hospitals)}")
    
    # Get batch
    batch = hospitals[:BATCH_SIZE]
    print(f"Generating template for {len(batch)} hospitals")
    print()
    
    # Generate YAML
    yaml_content = [
        "# NEXT BATCH TEMPLATE",
        f"# Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        "#",
        "# Instructions:",
        "#   1. Fill in the 'url' field with the leadership page URL",
        "#   2. Update 'pattern' to match the HTML structure (see pattern guide)",
        "#   3. Set 'expected_executives' based on website inspection",
        "#   4. Update 'status' from 'needs_configuration' to 'needs_testing' once URL added",
        "#   5. Add any notes about the hospital's structure",
        "#   6. Once tested and working, copy to enhanced_hospitals.yaml and mark done='y' in Excel",
        "#",
        "# Common patterns:",
        "#   - h2_name_h3_title: Names in <h2>, titles in <h3>",
        "#   - combined_h2: Name and title together with separator",
        "#   - div_classes: CSS class-based extraction",
        "#   - table_rows: Table structure",
        "#   - list_items: List with separators",
        "#",
        "# See PATTERN_REGISTRY.md for all 13 patterns and examples",
        "",
        "hospitals:"
    ]
    
    # Add each hospital
    for h in batch:
        yaml_content.append(generate_yaml_entry(h['FAC'], h['Hospital']))
    
    # Add summary
    yaml_content.extend([
        "",
        "# ============================================================================",
        "# BATCH SUMMARY",
        "# ============================================================================",
        f"# Total active hospitals in this batch: {len(batch)}",
        f"# Remaining after this batch: {len(hospitals) - len(batch)}",
        "#",
        "# WORKFLOW:",
        "# 1. For each hospital, find the leadership page URL",
        "# 2. Inspect the page structure in browser DevTools",
        "# 3. Update pattern, expected_executives, and html_structure fields",
        "# 4. Change status to 'needs_testing'",
        "# 5. Copy entry to enhanced_hospitals.yaml",
        "# 6. Test with: quick_test(FAC)",
        "# 7. If successful, change status to 'ok' or 'configured'",
        "# 8. Mark as done='y' in the LeadershipURLYAMLNotes.xlsx file",
        "#",
        "# Remember: Work directly in enhanced_hospitals.yaml for testing!",
        "# This file is just a reference/tracking list.",
        "# ============================================================================",
        ""
    ])
    
    # Write file
    yaml_text = '\n'.join(yaml_content)
    with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
        f.write(yaml_text)
    
    print("=" * 70)
    print(f"✓ Generated {OUTPUT_FILE}")
    print("=" * 70)
    print()
    print("NEXT STEPS:")
    print("1. Review the generated YAML file")
    print("2. For each hospital, find the leadership page URL")
    print("3. Copy entries to enhanced_hospitals.yaml for testing")
    print("4. Use quick_test(FAC) to test each hospital")
    print("5. Mark done='y' in Excel when complete")
    print()
    print(f"Hospitals in this batch: {len(batch)}")
    for i, h in enumerate(batch, 1):
        fac = format_fac(h['FAC'])
        print(f"  {i:2d}. FAC-{fac}: {h['Hospital']}")

if __name__ == '__main__':
    main()